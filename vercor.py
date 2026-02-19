"""
AUTOMAÇÃO BIOCROMA (HEADLESS + MULTIUSUÁRIO + EXCEL IDs/LOG + DOWNLOAD POR ID)

O que faz:
- Pergunta no CMD qual usuário (1/2/3...) do arquivo usuarios.json
- Abre Chrome INVISÍVEL (headless)
- Faz login automático (sem salvar senha no Chrome)
- Vai para /app/pedidos-paternidade
- Para cada ID na aba "IDs" do controle.xlsx:
    - Clica no filtro (sempre)
    - Digita a requisição e pesquisa (Enter)
    - Espera até 2s para aparecer o status
    - Se status == "Liberado": baixa o PDF e salva como {ID}.pdf
    - Loga em Excel (aba "Log") com cores
    - Remove da aba "IDs" os que concluíram (LIBERADO_BAIXADO ou JA_EXISTE)
- Salva a planilha no final

Arquivos necessários na MESMA pasta do script:
1) usuarios.json  (credenciais por usuário)
2) controle.xlsx  (ou ele cria se não existir)

Instalar deps:
    python -m pip install selenium requests openpyxl
"""

import json
import os
import time
from datetime import datetime
import requests

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


# ================== CONFIG ==================
USUARIOS_JSON = "usuarios.json"

SITE_URL = "https://biocroma.neovita.net.br"
LOGIN_URL_PADRAO = "https://biocroma.neovita.net.br/login"
PAGINA_PEDIDOS = "https://biocroma.neovita.net.br/app/pedidos-paternidade"

ARQUIVO_XLSX = "controle.xlsx"
ABA_IDS = "IDs"
ABA_LOG = "Log"

DOWNLOAD_DIR = r"C:\Automacao\PDFs"

HEADLESS = True
WAIT = 15

STATUS_TIMEOUT = 2        # esperar status até 2s
DOWNLOAD_TIMEOUT = 2      # esperar link até 2s
RETRY_PESQUISA = 1        # tentar mais 1 vez se status não aparecer
# ===========================================


# ================== SELETORES ==================
# Login (confirmado por você)
LOGIN_EMAIL = (By.CSS_SELECTOR, "input#email")
LOGIN_SENHA = (By.CSS_SELECTOR, "input#password")
LOGIN_BTN = (By.CSS_SELECTOR, "form button[type='submit']")

# Menu e página
MENU_PEDIDOS = (By.CSS_SELECTOR, 'a[href="https://biocroma.neovita.net.br/app/pedidos-paternidade"]')

# Filtro / modal / busca
FILTRO_BTN = (By.CSS_SELECTOR, "a.modal-trigger.btn-acao")
MODAL = (By.ID, "modal")
ID_INPUT = (By.NAME, "requisicao")

# Status + download
STATUS_BADGE = (By.CSS_SELECTOR, 'td[data-label="Situação:"] span.new.badge')
DOWNLOAD_LINK = (By.CSS_SELECTOR, "a.btn-download-exame")
# ===============================================


# ================== UTIL: USUÁRIOS ==================
def carregar_usuarios(path: str) -> dict:
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Não achei {path}.\n"
            "Crie um arquivo usuarios.json com este formato:\n"
            '{\n'
            '  "1": {"nome":"Você (ADM)","login_url":"https://.../login","email":"...","senha":"..."},\n'
            '  "2": {"nome":"Parceiro","login_url":"https://.../login","email":"...","senha":"..."}\n'
            '}\n'
        )
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def escolher_usuario(usuarios: dict) -> dict:
    print("\nSelecione o usuário:")
    for k, u in usuarios.items():
        print(f"  {k}) {u.get('nome','(sem nome)')} — {u.get('email','')}")
    escolha = input("Digite 1, 2, 3... e Enter: ").strip()

    if escolha not in usuarios:
        raise ValueError("Opção inválida. Escolha um número existente no usuarios.json.")

    u = usuarios[escolha]
    u.setdefault("login_url", LOGIN_URL_PADRAO)

    if not u.get("email") or not u.get("senha") or not u.get("login_url"):
        raise ValueError("Usuário escolhido está sem email/senha/login_url no usuarios.json.")

    return u


# ================== UTIL: EXCEL ==================
def criar_planilha_se_nao_existir(path: str) -> None:
    if os.path.exists(path):
        return

    wb = Workbook()
    ws_ids = wb.active
    ws_ids.title = ABA_IDS
    ws_ids["A1"] = "ID"

    ws_log = wb.create_sheet(ABA_LOG)
    ws_log.append(["Data/Hora", "ID", "Status", "Arquivo", "Obs"])

    wb.save(path)


def abrir_planilha(path: str):
    criar_planilha_se_nao_existir(path)
    wb = load_workbook(path)

    if ABA_IDS not in wb.sheetnames:
        ws_ids = wb.create_sheet(ABA_IDS)
        ws_ids["A1"] = "ID"

    if ABA_LOG not in wb.sheetnames:
        ws_log = wb.create_sheet(ABA_LOG)
        ws_log.append(["Data/Hora", "ID", "Status", "Arquivo", "Obs"])

    return wb


def ler_ids_da_planilha(wb) -> list[str]:
    ws = wb[ABA_IDS]
    ids = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        v = row[0]
        if v is None:
            continue
        s = str(v).strip()
        if s:
            ids.append(s)
    return ids


def aplicar_estilo_log(ws_log):
    header_fill = PatternFill("solid", fgColor="1F4E79")  # azul escuro
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, 6):
        cell = ws_log.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [20, 14, 20, 45, 35]
    for i, w in enumerate(widths, start=1):
        ws_log.column_dimensions[get_column_letter(i)].width = w


def cor_por_status(status: str) -> PatternFill:
    s = (status or "").upper()
    if s in {"LIBERADO_BAIXADO", "JA_EXISTE"}:
        return PatternFill("solid", fgColor="C6EFCE")  # verde claro
    if s in {"NAO_LIBERADO"}:
        return PatternFill("solid", fgColor="FFEB9C")  # amarelo
    if s in {"TIMEOUT_STATUS", "LIBERADO_SEM_LINK"}:
        return PatternFill("solid", fgColor="BDD7EE")  # azul claro
    if s in {"ERRO"}:
        return PatternFill("solid", fgColor="FFC7CE")  # vermelho claro
    return PatternFill("solid", fgColor="FFFFFF")


def adicionar_linha_log(wb, dt: str, doc_id: str, status: str, arquivo: str = "", obs: str = ""):
    ws = wb[ABA_LOG]
    if ws.max_row == 1:
        aplicar_estilo_log(ws)

    ws.append([dt, doc_id, status, arquivo, obs])

    r = ws.max_row
    fill = cor_por_status(status)
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=c).alignment = Alignment(vertical="center")


def remover_ids_concluidos(wb, ids_concluidos: set[str]):
    ws = wb[ABA_IDS]
    for row in range(ws.max_row, 1, -1):
        v = ws.cell(row=row, column=1).value
        if v is None:
            continue
        s = str(v).strip()
        if s in ids_concluidos:
            ws.delete_rows(row, 1)


# ================== UTIL: SELENIUM ==================
def criar_driver(download_dir: str, headless: bool) -> webdriver.Chrome:
    os.makedirs(download_dir, exist_ok=True)

    options = Options()

    if headless:
        options.add_argument("--headless=new")

    # estabilidade headless
    options.add_argument("--window-size=1200,900")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    # não sugerir salvar senha / credenciais
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True,
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
    }
    options.add_experimental_option("prefs", prefs)

    return webdriver.Chrome(options=options)


def fazer_login(driver: webdriver.Chrome, wait: WebDriverWait, login_url: str, email: str, senha: str) -> None:
    driver.get(login_url)

    email_inp = wait.until(EC.presence_of_element_located(LOGIN_EMAIL))
    senha_inp = wait.until(EC.presence_of_element_located(LOGIN_SENHA))

    email_inp.clear()
    email_inp.send_keys(email)

    senha_inp.clear()
    senha_inp.send_keys(senha)

    wait.until(EC.element_to_be_clickable(LOGIN_BTN)).click()
    time.sleep(1.0)  # pequena pausa para sessão firmar


def ir_para_pedidos(driver: webdriver.Chrome, wait: WebDriverWait) -> None:
    driver.get(PAGINA_PEDIDOS)
    time.sleep(0.5)

    # se por algum motivo não abriu direto, tenta pelo menu
    if not driver.current_url.startswith(PAGINA_PEDIDOS):
        try:
            wait.until(EC.element_to_be_clickable(MENU_PEDIDOS)).click()
            wait.until(lambda d: d.current_url.startswith(PAGINA_PEDIDOS))
        except Exception:
            driver.get(PAGINA_PEDIDOS)


def abrir_filtro_sempre(wait: WebDriverWait, driver: webdriver.Chrome) -> None:
    btn = wait.until(EC.element_to_be_clickable(FILTRO_BTN))
    try:
        btn.click()
    except Exception:
        driver.execute_script("arguments[0].click();", btn)

    wait.until(EC.visibility_of_element_located(MODAL))
    time.sleep(0.1)


def esperar_status(driver: webdriver.Chrome, timeout=2) -> str:
    fim = time.time() + timeout
    while time.time() < fim:
        els = driver.find_elements(*STATUS_BADGE)
        if els:
            caption = (els[0].get_attribute("data-badge-caption") or "").strip()
            if caption:
                return caption
        time.sleep(0.1)
    return ""


# ================== DOWNLOAD ==================
def cookies_para_requests(driver: webdriver.Chrome) -> requests.Session:
    s = requests.Session()
    for c in driver.get_cookies():
        s.cookies.set(c["name"], c["value"])
    return s


def baixar_pdf(sess: requests.Session, url: str, out_path: str) -> None:
    with sess.get(url, stream=True, timeout=60) as r:
        r.raise_for_status()
        with open(out_path, "wb") as f:
            for chunk in r.iter_content(1024 * 128):
                if chunk:
                    f.write(chunk)


# ================== MAIN ==================
def executar():
    usuarios = carregar_usuarios(USUARIOS_JSON)
    user = escolher_usuario(usuarios)

    wb = abrir_planilha(ARQUIVO_XLSX)
    ids = ler_ids_da_planilha(wb)

    if not ids:
        print("⚠️ Nenhum ID na aba 'IDs' do controle.xlsx.")
        return

    ids_concluidos = set()

    driver = criar_driver(DOWNLOAD_DIR, headless=HEADLESS)
    wait = WebDriverWait(driver, WAIT)

    try:
        print("🔐 Fazendo login (invisível)...")
        fazer_login(driver, wait, user["login_url"], user["email"], user["senha"])

        print("📄 Indo para a página de pedidos...")
        ir_para_pedidos(driver, wait)

        for i, doc_id in enumerate(ids, 1):
            dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print(f"\n[{i}/{len(ids)}] ID: {doc_id}")

            try:
                # filtro SEMPRE
                abrir_filtro_sempre(wait, driver)

                # pesquisar
                campo = wait.until(EC.element_to_be_clickable(ID_INPUT))
                campo.clear()
                campo.send_keys(doc_id)
                campo.send_keys("\n")

                status = esperar_status(driver, timeout=STATUS_TIMEOUT)

                if (not status) and RETRY_PESQUISA > 0:
                    # retry
                    abrir_filtro_sempre(wait, driver)
                    campo = wait.until(EC.element_to_be_clickable(ID_INPUT))
                    campo.clear()
                    campo.send_keys(doc_id)
                    campo.send_keys("\n")
                    status = esperar_status(driver, timeout=STATUS_TIMEOUT)

                if not status:
                    print("⚠️ Status não apareceu (timeout).")
                    adicionar_linha_log(wb, dt, doc_id, "TIMEOUT_STATUS", "", "status não apareceu em 2s")
                    continue

                print("Status:", status)

                if status != "Liberado":
                    adicionar_linha_log(wb, dt, doc_id, "NAO_LIBERADO", "", status)
                    continue

                # liberado -> download
                try:
                    link = WebDriverWait(driver, DOWNLOAD_TIMEOUT).until(
                        EC.presence_of_element_located(DOWNLOAD_LINK)
                    )
                except TimeoutException:
                    adicionar_linha_log(wb, dt, doc_id, "LIBERADO_SEM_LINK", "", "download link não apareceu em 2s")
                    continue

                url_pdf = link.get_attribute("href") or ""
                if not url_pdf:
                    adicionar_linha_log(wb, dt, doc_id, "LIBERADO_SEM_LINK", "", "href vazio")
                    continue

                out_path = os.path.join(DOWNLOAD_DIR, f"{doc_id}.pdf")

                if os.path.exists(out_path) and os.path.getsize(out_path) > 0:
                    adicionar_linha_log(wb, dt, doc_id, "JA_EXISTE", out_path, "")
                    ids_concluidos.add(doc_id)
                    continue

                sess = cookies_para_requests(driver)
                baixar_pdf(sess, url_pdf, out_path)

                adicionar_linha_log(wb, dt, doc_id, "LIBERADO_BAIXADO", out_path, "")
                ids_concluidos.add(doc_id)

            except Exception as e:
                adicionar_linha_log(wb, dt, doc_id, "ERRO", "", str(e))

        # remove os concluídos da aba IDs
        if ids_concluidos:
            remover_ids_concluidos(wb, ids_concluidos)

        wb.save(ARQUIVO_XLSX)
        print(f"\n✅ Finalizado. Planilha atualizada: {ARQUIVO_XLSX}")
        print(f"📁 Downloads em: {DOWNLOAD_DIR}")

    finally:
        driver.quit()


if __name__ == "__main__":
    executar()